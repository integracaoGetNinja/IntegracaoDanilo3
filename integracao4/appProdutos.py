import requests
from time import sleep
import tkinter as tk
from tkinter import messagebox
import webbrowser
import openpyxl
from pymongo import MongoClient
import threading
from tkinter import ttk
import os
from tabulate import tabulate

# pyinstaller --onefile --add-data "requirements.txt;." integracao4/appProdutos.py

client = MongoClient('mongodb+srv://integracaoDanilo4:8iz8NbjUfoUSxAQG@cluterb.ypmgnks.mongodb.net/')
db = client["credencialIntegracao4Danilo"]
col_bling = db["col_bling"]

TEMPO_ESPERA = 0.5


def list_more_sells():
    data_inicial = entry_data_inicial.get()
    data_final = entry_data_final.get()

    token = col_bling.find_one({"_id": 0}).get('token')

    largura_janela_principal = root.winfo_width()
    altura_janela_principal = root.winfo_height()

    largura_tela = root.winfo_screenwidth()
    altura_tela = root.winfo_screenheight()

    x = (largura_tela - largura_janela_principal) // 2
    y = (altura_tela - altura_janela_principal) // 2

    janela_de_carregamento = tk.Toplevel(root)
    janela_de_carregamento.title("Carregando...")
    janela_de_carregamento.geometry("300x100+{}+{}".format(x, y))
    janela_de_carregamento.protocol("WM_DELETE_WINDOW", lambda: None)

    progresso = ttk.Progressbar(janela_de_carregamento, orient="horizontal", length=250, mode="indeterminate")
    progresso.pack(pady=20)
    progresso.start()

    payload = {}
    headers = {
        'Accept': 'application/json',
        'Authorization': f'Bearer {token}',
        'Cookie': 'PHPSESSID=1a06ulj1eh9g3arn5hqjlsn8gq'
    }

    notasFiscais = []

    pagina = 1

    while True:
        url = f"https://www.bling.com.br/Api/v3/nfe?pagina={pagina}&dataEmissaoInicial={data_inicial}&dataEmissaoFinal={data_final}"
        response = requests.request("GET", url, headers=headers, data=payload)

        print(url)
        print(response.status_code)

        if not response.json().get('data'):
            break

        notasFiscais += response.json().get('data')
        pagina += 1
        print(f"Capturando pagina {pagina}\n{len(notasFiscais)} notas ficais encontradas.")
        sleep(TEMPO_ESPERA)

    # print(token)
    # print(response.status_code)

    listProdutos = []

    for notaFiscal in notasFiscais:
        # print(notaFiscal)
        url = f"https://www.bling.com.br/Api/v3/pedidos/vendas?numero={notaFiscal.get('numero')}"
        try:
            responseGetId = requests.request("GET", url, headers=headers, data=payload)

            idPedido = responseGetId.json().get('data')[0].get('id')

            url = f"https://www.bling.com.br/Api/v3/pedidos/vendas/{idPedido}"
            responseItens = requests.request("GET", url, headers=headers, data=payload).json().get('data')

            for item in responseItens.get('itens'):

                # faça um switch
                idloja = str(responseItens.get('loja').get('id'))
                canal_de_venda = idloja
                if idloja == "203768813":
                    canal_de_venda = 'Site'
                elif idloja == "204527966":
                    canal_de_venda = 'Shopee'
                elif idloja == '204525390':
                    canal_de_venda = 'Mercado Livre'
                elif idloja == '203890758':
                    canal_de_venda = 'Carrefour'
                elif idloja == '204540954':
                    canal_de_venda = 'B2W'
                elif idloja == '204525390':
                    canal_de_venda = 'Matriz'
                elif idloja == '204517821':
                    canal_de_venda = 'Amazon'
                elif idloja == '204517821':
                    canal_de_venda = 'Olist'

                produto = {
                    "nomeProduto": item.get('descricao'),
                    "quantidade": item.get('quantidade'),
                    "valor": item.get('comissao').get('base'),
                    "canalVenda": canal_de_venda
                }

                listProdutos = verifyAddedInListProdutos(listProdutos, produto)
                nl = sorted(listProdutos, key=lambda x: x['quantidade'], reverse=True)
                os.system("cls")
                tabela = tabulate(nl, headers='keys',
                                  tablefmt="fancy_grid")
                print(tabela)

            sleep(TEMPO_ESPERA)

        except Exception as err:
            print(err)

    print("listProdutos", len(listProdutos))
    nl = sorted(listProdutos, key=lambda x: x['quantidade'], reverse=True)

    nome_planilha = f'melhores_produtos_{data_inicial.replace("/", "")}_ate_{data_final.replace("/", "")}.xlsx'

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet['A1'] = 'Nome do Produto'
    sheet['B1'] = 'Quantidade'
    sheet['C1'] = 'Valor'
    sheet['D1'] = 'Canal de Venda'

    for index, item in enumerate(nl, start=2):
        sheet.cell(row=index, column=1, value=item['nomeProduto'])
        sheet.cell(row=index, column=2, value=item['quantidade'])
        sheet.cell(row=index, column=3, value=f'R$ {item["valor"]}')
        sheet.cell(row=index, column=4, value=item['canalVenda'])

    # Salvar a planilha

    if response.status_code == 401:
        abrir_credenciamento()
    elif response.status_code == 200:
        workbook.save(nome_planilha)
        messagebox.showinfo("Sucesso!", f"Arquivo salvo em {nome_planilha}")

    janela_de_carregamento.destroy()

    # return listProdutos


def verifyAddedInListProdutos(listProducts, product):
    inList = False
    for i in listProducts:
        if i.get('nomeProduto') == product.get('nomeProduto'):
            i['quantidade'] += product.get('quantidade')
            i['valor'] += product.get('valor')
            inList = True
            # print(i)
            break

    if not inList:
        listProducts.append(product)
        # print(product)

    return listProducts


def abrir_credenciamento():
    webbrowser.open(
        "https://www.bling.com.br/OAuth2/views/authorization.php?response_type=code&client_id=f2e78a693a0f7b9d17b85e08341c6553978bb680&state=d67ea464dbf4817a262970ed215edc11&scopes=98308+98309+98310+98313+98314+98619+101584+104142+104163+507943+575904+5990556+6631498+106168710+182224097+199272829+200802821+220621674+318257547+318257548+318257550+318257553+318257555+318257556+318257559+318257561+318257562+318257563+318257565+318257568+318257570+318257573+318257576+318257577+318257580+318257583+333936575+363921589+363921590+363921591+363921592+363921598+363921599+363953167+363953556+363953706+791588404+875116881+875116885+1649295804+1780272711+1869535257+5862218180+6239411327+6239420709+13645012976+13645012997+13645012998")


def obter_datas():
    thread = threading.Thread(target=list_more_sells)
    thread.start()


# Criar uma janela
root = tk.Tk()
root.title("Seleção de Datas")

# Definir a largura e altura da janela e centralizá-la na tela
largura = 400
altura = 300
x = (root.winfo_screenwidth() - largura) // 2
y = (root.winfo_screenheight() - altura) // 2
root.geometry(f"{largura}x{altura}+{x}+{y}")

# Criar rótulos e campos de entrada para as datas
label_data_inicial = tk.Label(root, text="Padrão Dia/Mês/Ano")
label_data_inicial.pack(pady=10)

label_data_inicial = tk.Label(root, text="Data Inicial:")
label_data_inicial.pack()
entry_data_inicial = tk.Entry(root)
entry_data_inicial.pack(pady=10)

label_data_final = tk.Label(root, text="Data Final:")
label_data_final.pack()
entry_data_final = tk.Entry(root)
entry_data_final.pack(pady=10)

# Criar um botão para confirmar as datas
botao_pronto = tk.Button(root, text="Pronto", command=obter_datas)
botao_pronto.pack(pady=20)

botao_google = tk.Button(root, text="Fazer Login", command=abrir_credenciamento)
botao_google.pack(pady=20)

# Iniciar o loop principal da interface gráfica
root.mainloop()
